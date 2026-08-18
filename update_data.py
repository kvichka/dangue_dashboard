#!/usr/bin/env python3
"""
update_data.py — rebuilds Data/dengue-data.json from Dengue_Master_Data_Entry.xlsx,
for the GitHub-hosted version of the dashboard (index.html + Data/ folder).

EASIEST WAY TO RUN IT:
  Windows  -> double-click update_data.bat
  Mac      -> double-click update_data.command (or run: python3 update_data.py)
  Anyone   -> open a terminal in this folder and run:  python3 update_data.py

It needs two files in the SAME folder as this script:
    - Dengue_Master_Data_Entry.xlsx   (your data — edit this every week)
    - static_assets.json              (rarely-changing reference data — don't touch)

It writes: Data/dengue-data.json  (creating the Data folder if it doesn't exist)

index.html never needs to change for a routine data update — only this JSON
file does. After running this script, just upload the new Data/dengue-data.json
to GitHub (replacing the old one) and commit; Pages redeploys automatically.

WHAT THIS SCRIPT CAN'T UPDATE (and why):
  "Monthly Case Composition by Severity" (national) needs a DF/DHF/DSS
  breakdown for every month nationally. That level of detail only exists in
  the original 800,000-row historical line-list, not in the simplified
  Master workbook (which only tracks total cases/deaths per province-month).
  It stays frozen at whatever it was the last time Claude rebuilt the
  dashboard from the full source files. Everything else — case counts, IR,
  CFR, the risk map, and age-group breakdowns — updates fully from your
  Excel file.
"""
import json, sys, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl is not installed yet — installing it now, one time only...")
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=False)
    try:
        import openpyxl
    except ImportError:
        print("\nCould not install openpyxl automatically.")
        print("Please run this yourself, then try again:  pip install openpyxl")
        input("\nPress Enter to close...")
        sys.exit(1)

HERE = Path(__file__).parent
EXCEL_PATH = HERE / "Dengue_Master_Data_Entry.xlsx"
STATIC_PATH = HERE / "static_assets.json"
OUTPUT_DIR = HERE / "Data"
OUTPUT_PATH = OUTPUT_DIR / "dengue-data.json"

MONTH_ABBR = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
MONTH_TO_NUM = {m: i+1 for i, m in enumerate(MONTH_ABBR)}

def log(msg):
    print(f"  {msg}")

def fail(msg):
    print(f"\nERROR: {msg}")
    input("\nPress Enter to close...")
    sys.exit(1)

print("=" * 60)
print("Dengue Dashboard — local rebuild")
print("=" * 60)

for p in [EXCEL_PATH, STATIC_PATH]:
    if not p.exists():
        fail(f"Required file not found: {p.name}\nMake sure it's in the same folder as this script.")

log(f"Reading {EXCEL_PATH.name} ...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
for needed in ["National_Monthly_Entry", "Population_by_Province_Year", "Lookup_Lists"]:
    if needed not in wb.sheetnames:
        fail(f"Sheet '{needed}' not found in the workbook. Did you rename or delete a sheet?")

static = json.loads(STATIC_PATH.read_text(encoding="utf-8"))
log("Loaded static reference data (map, geo boundaries, facilities, historical severity mix).")

# ---- Province name mapping (Full Name <-> DHF bulletin abbreviation) ----
ws_lookup = wb["Lookup_Lists"]
name_map = {}       # abbrev -> full
full_to_abbrev = {} # full -> abbrev
r = 5
while True:
    full = ws_lookup.cell(row=r, column=1).value
    abbrev = ws_lookup.cell(row=r, column=2).value
    if not full:
        break
    name_map[abbrev] = full
    full_to_abbrev[full] = abbrev
    r += 1
log(f"Loaded {len(name_map)} provinces from Lookup_Lists.")

# ---- Population_by_Province_Year ----
ws_pop = wb["Population_by_Province_Year"]
pop_years = []
c = 2
while True:
    v = ws_pop.cell(row=4, column=c).value
    if v is None:
        break
    pop_years.append(int(str(v).split()[0]))  # handles "2026 (YTD)"-style just in case
    c += 1
full_population = {}
r = 5
while True:
    prov = ws_pop.cell(row=r, column=1).value
    if not prov:
        break
    full_population[prov] = {}
    for i, y in enumerate(pop_years):
        val = ws_pop.cell(row=r, column=2+i).value
        full_population[prov][str(y)] = val
    r += 1
log(f"Loaded population for {len(full_population)} provinces, years {pop_years[0]}-{pop_years[-1]}.")

# ---- National_Monthly_Entry: read every real row (province, year, month, cases, deaths) ----
# Starts at row 5 (not a hardcoded "row 6") and skips only the row whose Province is literally
# "EXAMPLE" — text-based, not font-based, since cell styling (italic) turned out to be an
# unreliable marker (real data rows can end up italic too, e.g. from copy/paste in Excel).
ws_nat = wb["National_Monthly_Entry"]
national_rows = []
r = 5
while True:
    prov_cell = ws_nat.cell(row=r, column=1)
    prov = prov_cell.value
    if prov is None:
        blank_run = 0
        rr = r
        while ws_nat.cell(row=rr, column=1).value is None and blank_run < 20:
            blank_run += 1
            rr += 1
        if blank_run >= 20:
            break
        r += 1
        continue
    is_example = isinstance(prov, str) and prov.strip().upper() == "EXAMPLE"
    year = ws_nat.cell(row=r, column=2).value
    month = ws_nat.cell(row=r, column=3).value
    cases = ws_nat.cell(row=r, column=4).value or 0
    deaths = ws_nat.cell(row=r, column=5).value or 0
    if prov and year and month and not is_example:
        national_rows.append({"province": prov, "year": int(year), "month": month, "cases": int(cases), "deaths": int(deaths)})
    r += 1
log(f"Loaded {len(national_rows)} national province-month rows.")

# =========================================================
# Rebuild every "dynamic" bundle field from the sheets above
# =========================================================
import statistics

CURRENT_YEAR = max(r["year"] for r in national_rows) if national_rows else datetime.date.today().year
HIST_YEARS = [y for y in range(2018, CURRENT_YEAR)]
log(f"Treating {CURRENT_YEAR} as the current (year-to-date) year; {HIST_YEARS[0]}-{HIST_YEARS[-1]} as history.")

# ---- multiyear.province_series: cases & deaths by province + year ----
province_series = {}
for row in national_rows:
    prov = row["province"]
    province_series.setdefault(prov, {"cases": {}, "deaths": {}})
    y = str(row["year"])
    province_series[prov]["cases"][y] = province_series[prov]["cases"].get(y, 0) + row["cases"]
    province_series[prov]["deaths"][y] = province_series[prov]["deaths"].get(y, 0) + row["deaths"]

all_years_present = sorted({row["year"] for row in national_rows})
national_yearly = []
for y in all_years_present:
    c = sum(r["cases"] for r in national_rows if r["year"] == y)
    d = sum(r["deaths"] for r in national_rows if r["year"] == y)
    national_yearly.append({"year": y, "cases": c, "deaths": d})

multiyear = {"years": all_years_present, "province_series": province_series,
             "national_yearly": national_yearly, "name_map": name_map}

# ---- monthly: national + per-province monthly cases/deaths, historical years only ----
national_monthly_cases, national_monthly_deaths = {}, {}
province_monthly_cases, province_monthly_deaths = {}, {}
for row in national_rows:
    if row["year"] not in HIST_YEARS:
        continue
    y, m, prov = str(row["year"]), str(MONTH_TO_NUM[row["month"]]), row["province"]
    national_monthly_cases.setdefault(y, {}); national_monthly_cases[y][m] = national_monthly_cases[y].get(m, 0) + row["cases"]
    national_monthly_deaths.setdefault(y, {}); national_monthly_deaths[y][m] = national_monthly_deaths[y].get(m, 0) + row["deaths"]
    province_monthly_cases.setdefault(prov, {}).setdefault(y, {})
    province_monthly_cases[prov][y][m] = province_monthly_cases[prov][y].get(m, 0) + row["cases"]
    province_monthly_deaths.setdefault(prov, {}).setdefault(y, {})
    province_monthly_deaths[prov][y][m] = province_monthly_deaths[prov][y].get(m, 0) + row["deaths"]
monthly = {"national_monthly_cases": national_monthly_cases, "national_monthly_deaths": national_monthly_deaths,
           "province_monthly_cases": province_monthly_cases, "province_monthly_deaths": province_monthly_deaths}

# ---- total_cases_2026 (current-year monthly detail, by province abbreviation) ----
current_year_rows = [r for r in national_rows if r["year"] == CURRENT_YEAR]
provinces_2026 = []
for prov_full in sorted({r["province"] for r in current_year_rows}):
    abbrev = full_to_abbrev.get(prov_full, prov_full)
    pop2026 = full_population.get(prov_full, {}).get(str(CURRENT_YEAR))
    monthly_data = {m: {"cases": 0, "deaths": 0} for m in MONTH_ABBR}
    for r in current_year_rows:
        if r["province"] == prov_full:
            monthly_data[r["month"]] = {"cases": r["cases"], "deaths": r["deaths"]}
    provinces_2026.append({"province": abbrev, "population2026": pop2026, "monthly": monthly_data,
                            "total_cases": sum(v["cases"] for v in monthly_data.values()),
                            "total_deaths": sum(v["deaths"] for v in monthly_data.values())})
national_total_2026 = {
    "population2026": sum(full_population.get(p, {}).get(str(CURRENT_YEAR)) or 0 for p in full_population),
    "monthly": {m: {"cases": sum(pp["monthly"][m]["cases"] for pp in provinces_2026),
                    "deaths": sum(pp["monthly"][m]["deaths"] for pp in provinces_2026)} for m in MONTH_ABBR},
}
national_total_2026["total_cases"] = sum(v["cases"] for v in national_total_2026["monthly"].values())
national_total_2026["total_deaths"] = sum(v["deaths"] for v in national_total_2026["monthly"].values())
total_cases_2026 = {"provinces_2026": provinces_2026, "national_total_2026": national_total_2026}

# ---- national: province list with cases/deaths YTD (same-period comparison) ----
# NOTE: the Master workbook only has monthly granularity, so "same period" here means
# "January through the latest reported month" for both years — not the exact ISO week
# comparison the original DHF bulletin used. Documented in the dashboard's own caption.
latest_month_num = max(MONTH_TO_NUM[r["month"]] for r in current_year_rows) if current_year_rows else 12
months_ytd = set(MONTH_ABBR[:latest_month_num])
prev_year = CURRENT_YEAR - 1
national_provinces = []
for abbrev, full in name_map.items():
    cases_now = sum(r["cases"] for r in national_rows if r["province"] == full and r["year"] == CURRENT_YEAR and r["month"] in months_ytd)
    cases_prev = sum(r["cases"] for r in national_rows if r["province"] == full and r["year"] == prev_year and r["month"] in months_ytd)
    deaths_now = sum(r["deaths"] for r in national_rows if r["province"] == full and r["year"] == CURRENT_YEAR and r["month"] in months_ytd)
    deaths_prev = sum(r["deaths"] for r in national_rows if r["province"] == full and r["year"] == prev_year and r["month"] in months_ytd)
    cfr_now = round(deaths_now/cases_now*100, 3) if cases_now else 0
    cfr_prev = round(deaths_prev/cases_prev*100, 3) if cases_prev else 0
    national_provinces.append({"province": abbrev, "cases2026": cases_now, "cases2025": cases_prev,
                                "deaths2026": deaths_now, "deaths2025": deaths_prev,
                                "cfr2026": cfr_now, "cfr2025": cfr_prev})
tot_c, tot_c5, tot_d, tot_d5 = (sum(p[k] for p in national_provinces) for k in ("cases2026","cases2025","deaths2026","deaths2025"))

# ---- age_cases / diag_age: read from National_AgeGroup_Entry (long format: one row per
# age group per year), pivot into current-year vs prior-year for the dashboard's charts.
# A Province column exists for future use (per-province age breakdowns) — for now, only
# "All Cambodia" rows feed these national charts; province-specific rows are read but ignored
# until a province-level age chart exists to consume them.
# Falls back to the frozen static snapshot for older workbooks built before this sheet existed. ----
if "National_AgeGroup_Entry" in wb.sheetnames:
    ws_age = wb["National_AgeGroup_Entry"]
    age_long = []  # (age_group, year, cases, deaths, DF, DHF, DSS)
    r = 5
    while True:
        ag_cell = ws_age.cell(row=r, column=1)
        ag = ag_cell.value
        if ag is None:
            blank_run = 0
            rr = r
            while ws_age.cell(row=rr, column=1).value is None and blank_run < 20:
                blank_run += 1; rr += 1
            if blank_run >= 20:
                break
            r += 1
            continue
        yr = ws_age.cell(row=r, column=2).value
        province_val = ws_age.cell(row=r, column=9).value  # column I; absent in older workbooks
        is_national = province_val is None or str(province_val).strip() == "All Cambodia"
        # No example-row filter needed here: the pivot step below only keeps rows matching
        # CURRENT_YEAR or CURRENT_YEAR-1, so the example row's year (2024) is naturally excluded.
        if yr is not None and is_national:
            age_long.append({
                "age_group": ag, "year": int(yr),
                "cases": ws_age.cell(row=r, column=3).value or 0,
                "deaths": ws_age.cell(row=r, column=4).value or 0,
                "DF": ws_age.cell(row=r, column=6).value or 0,
                "DHF": ws_age.cell(row=r, column=7).value or 0,
                "DSS": ws_age.cell(row=r, column=8).value or 0,
            })
        r += 1
    age_groups_present = list(dict.fromkeys(x["age_group"] for x in age_long))  # preserve first-seen order
    prior_year = CURRENT_YEAR - 1
    age_cases, diag_age = [], []
    for ag in age_groups_present:
        cur = next((x for x in age_long if x["age_group"]==ag and x["year"]==CURRENT_YEAR), None)
        prev = next((x for x in age_long if x["age_group"]==ag and x["year"]==prior_year), None)
        c26 = cur["cases"] if cur else 0
        d26 = cur["deaths"] if cur else 0
        c25 = prev["cases"] if prev else 0
        d25 = prev["deaths"] if prev else 0
        age_cases.append({"age_group": ag, "cases2026": c26, "cases2025": c25,
                           "deaths2026": d26, "deaths2025": d25,
                           "cfr2026": round(d26/c26*100,3) if c26 else 0,
                           "cfr2025": round(d25/c25*100,3) if c25 else 0})
        diag_age.append({"age_group": ag, "DF": cur["DF"] if cur else 0,
                          "DHF": cur["DHF"] if cur else 0, "DSS": cur["DSS"] if cur else 0})
    log(f"Loaded age-group data for {len(age_cases)} age bands ({CURRENT_YEAR} vs {prior_year}) from National_AgeGroup_Entry.")
else:
    age_cases, diag_age = static["national_age_cases"], static["national_diag_age"]
    log("WARNING: National_AgeGroup_Entry sheet not found — using frozen age-group data instead.")

national = {"provinces": national_provinces,
            "total": {"cases2026": tot_c, "cases2025": tot_c5, "deaths2026": tot_d, "deaths2025": tot_d5,
                      "cfr2026": round(tot_d/tot_c*100,3) if tot_c else 0, "cfr2025": round(tot_d5/tot_c5*100,3) if tot_c5 else 0},
            "age_cases": age_cases, "diag_age": diag_age}

log(f"National YTD ({CURRENT_YEAR} through {MONTH_ABBR[latest_month_num-1]}): {tot_c} cases vs {tot_c5} same period {prev_year}.")

# ---- ir_heatmap: IR by province by year, using real population ----
ir_by_province_year = {}
GROWTH = 1.024
for prov_full in full_population:
    series = province_series.get(prov_full, {"cases": {}})
    ir_by_province_year[prov_full] = {}
    for y in range(min(HIST_YEARS[0], 2018), CURRENT_YEAR+1):
        cases = series["cases"].get(str(y), 0)
        pop = full_population[prov_full].get(str(y))
        if pop is None:
            # outside the loaded population sheet's range — extrapolate from nearest known year
            base_year = min(full_population[prov_full], key=lambda yy: abs(int(yy)-y))
            pop = full_population[prov_full][base_year] / (GROWTH ** (y - int(base_year)))
        ir_by_province_year[prov_full][str(y)] = round(cases/pop*100000, 1) if pop else 0
ir_heatmap = {"years": list(range(min(HIST_YEARS[0], 2018), CURRENT_YEAR+1)), "ir_by_province_year": ir_by_province_year}

# ---- epidemic_channel: national monthly mean/std across historical years ----
channel = []
for m_idx, m_abbr in enumerate(MONTH_ABBR, start=1):
    vals = [national_monthly_cases.get(str(y), {}).get(str(m_idx), 0) for y in HIST_YEARS]
    mean = statistics.mean(vals) if vals else 0
    std = statistics.stdev(vals) if len(vals) > 1 else 0
    channel.append({"month": m_abbr, "normal": round(mean,1), "alert": round(mean+std,1), "epidemic": round(mean+2*std,1)})
current_year_monthly = [None]*12
for r in current_year_rows:
    idx = MONTH_TO_NUM[r["month"]] - 1
    current_year_monthly[idx] = (current_year_monthly[idx] or 0) + r["cases"]
epidemic_channel = {"channel": channel, "current_year_2026": current_year_monthly}

log(f"Rebuilt national aggregates ({len(national_rows)} rows).")

# =========================================================
# Merge dynamic data with static assets, inject into the HTML template
# =========================================================
bundle = {
    "national": national,
    "map": static["map"],                                    # static — see note in file header
    "multiyear": multiyear,
    "monthly": monthly,
    "monthly_severity": static["monthly_severity"],           # static — see note in file header
    "province_monthly_severity": static["province_monthly_severity"],  # static
    "population": static["population"],                       # legacy field, kept for compatibility
    "total_cases_2026": total_cases_2026,
    "ir_heatmap": ir_heatmap,
    "key_facilities": static["key_facilities"],               # static
    "cambodia_geo": static["cambodia_geo"],                   # static
    "epidemic_channel": epidemic_channel,
    "full_population_2018_2026": full_population,
}

log("Writing data file ...")
OUTPUT_DIR.mkdir(exist_ok=True)
data_json = json.dumps(bundle, ensure_ascii=False)
OUTPUT_PATH.write_text(data_json, encoding="utf-8")
size_kb = OUTPUT_PATH.stat().st_size / 1024

print()
print("=" * 60)
print(f"DONE. Wrote {OUTPUT_PATH.relative_to(HERE)} ({size_kb:.0f} KB).")
print("=" * 60)
print()
print("Next steps:")
print("  1. Preview it locally first if you want: from this folder run")
print("       python3 -m http.server 8000")
print("     then open http://localhost:8000/ in a browser (double-clicking")
print("     index.html directly will NOT work for loading the data).")
print("  2. Upload Data/dengue-data.json to your GitHub repo, replacing the")
print("     old one (same filename, same Data/ folder), and commit.")
print("  3. GitHub Pages redeploys automatically within a minute or two —")
print("     no changes to index.html are needed for a routine data update.")
print()
print("Reminder: 'Monthly Case Composition by Severity' (national) is frozen at its last")
print("Claude-built values — the Master workbook doesn't carry national diagnosis-by-month")
print("detail. Everything else, including age-group breakdowns, reflects the workbook now.")
input("\nPress Enter to close...")
